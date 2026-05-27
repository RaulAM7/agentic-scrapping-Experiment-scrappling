from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUT_DIR = Path("04_outputs/skilland-ia-mujeres/cabildos_contactos_igualdad_social")
CSV_PATH = OUT_DIR / "directorio_cabildos_igualdad_social.csv"
XLSX_PATH = OUT_DIR / "directorio_cabildos_igualdad_social.xlsx"
MD_PATH = OUT_DIR / "directorio_cabildos_igualdad_social.md"
NOTES_PATH = OUT_DIR / "scraping_notes.md"


HEADERS = [
    "Isla",
    "Cabildo",
    "Area_Consejeria_Servicio",
    "Cargo_exacto",
    "Nombre_y_apellidos",
    "Partido_politico",
    "Email_personal_institucional",
    "Email_area_consejeria",
    "Email_generico_fallback",
    "Telefono",
    "URL_fuente_principal",
    "URL_fuente_secundaria",
    "Nivel_confianza",
    "Motivo_confianza",
    "Notas_outreach",
    "Contacto_principal_recomendado",
    "Comparativa_excel_inicial",
]


ROWS = [
    {
        "Isla": "Gran Canaria",
        "Cabildo": "Cabildo de Gran Canaria",
        "Area_Consejeria_Servicio": "Politica Social, Accesibilidad, Igualdad y Diversidad",
        "Cargo_exacto": "Vicepresidenta 3a; Consejera de Politica Social, Accesibilidad, Igualdad y Diversidad",
        "Nombre_y_apellidos": "Isabel Mena Alonso",
        "Partido_politico": "PSOE / Grupo Politico Socialista",
        "Email_personal_institucional": "Unknown",
        "Email_area_consejeria": "igualdadcabildo@grancanaria.com",
        "Email_generico_fallback": "oiac@grancanaria.com",
        "Telefono": "+34 928 219 421 / +34 928 219 494",
        "URL_fuente_principal": "https://cabildo.grancanaria.com/organigrama",
        "URL_fuente_secundaria": "https://igualdad.grancanaria.com/directorio",
        "Nivel_confianza": "Alto",
        "Motivo_confianza": "Organigrama oficial actualizado el 04/03/2026 confirma cargo y persona; directorio oficial de igualdad confirma email del area.",
        "Notas_outreach": "Contacto principal recomendado para propuesta de empoderamiento femenino en IA. Usar email del area y mencionar Consejeria de Politica Social/Igualdad.",
        "Contacto_principal_recomendado": "Si",
        "Comparativa_excel_inicial": "Confirmado; se mantiene responsable y se conserva email del area como via preferente.",
    },
    {
        "Isla": "Gran Canaria",
        "Cabildo": "Cabildo de Gran Canaria",
        "Area_Consejeria_Servicio": "Empleo y Desarrollo Local",
        "Cargo_exacto": "Consejero de Gobierno de Empleo y Desarrollo Local",
        "Nombre_y_apellidos": "Juan Diaz Sanchez",
        "Partido_politico": "PSOE / Grupo Politico Socialista",
        "Email_personal_institucional": "Unknown",
        "Email_area_consejeria": "Unknown",
        "Email_generico_fallback": "oiac@grancanaria.com",
        "Telefono": "+34 928 219 421 / +34 928 219 494",
        "URL_fuente_principal": "https://cabildo.grancanaria.com/-/juan-diaz-sanchez",
        "URL_fuente_secundaria": "https://cabildo.grancanaria.com/organigrama",
        "Nivel_confianza": "Medio",
        "Motivo_confianza": "Cargo confirmado en ficha y organigrama oficiales; no se localizo email directo ni email publico especifico del area.",
        "Notas_outreach": "Contacto secundario para empleabilidad. Enviar por OIAC solo si no se dispone de canal interno o contacto previo en Cabildo.",
        "Contacto_principal_recomendado": "No",
        "Comparativa_excel_inicial": "Confirmado cargo/persona; email generico del Excel no queda mejorado con fuente especifica de area.",
    },
    {
        "Isla": "Tenerife",
        "Cabildo": "Cabildo de Tenerife",
        "Area_Consejeria_Servicio": "Direccion Insular de Igualdad y Diversidad / Unidad Tecnica de Igualdad y Diversidad",
        "Cargo_exacto": "Directora Insular de Igualdad y Diversidad",
        "Nombre_y_apellidos": "Patricia Leon Perez",
        "Partido_politico": "Unknown",
        "Email_personal_institucional": "Unknown",
        "Email_area_consejeria": "igualdadydiversidad@tenerife.es",
        "Email_generico_fallback": "atencionciudadana@tenerife.es",
        "Telefono": "922 239 707 / 922 239 466",
        "URL_fuente_principal": "https://www.tenerife.es/contacto-unidad-de-igualdad",
        "URL_fuente_secundaria": "https://www.tenerife.es/w/el-cabildo-consolida-una-estrategia-integral-de-igualdad-con-impacto-real-en-toda-la-isla-que-llega-a-mas-de-450.000-personas",
        "Nivel_confianza": "Alto",
        "Motivo_confianza": "Fuente oficial confirma email y telefonos de la Unidad; noticias oficiales 2025-2026 identifican a Patricia Leon como directora insular que lidera la estrategia.",
        "Notas_outreach": "Mejor puerta de entrada para dossier de mujeres e IA en igualdad: la Direccion Insular gestiona Tenerife Violeta y proyectos medibles.",
        "Contacto_principal_recomendado": "Si",
        "Comparativa_excel_inicial": "Mejorado; el Excel ponia como contacto principal a Presidencia. Se propone contacto operativo mas cercano al programa.",
    },
    {
        "Isla": "Tenerife",
        "Cabildo": "Cabildo de Tenerife",
        "Area_Consejeria_Servicio": "Accion Social, Inclusion, Voluntariado y Participacion Ciudadana",
        "Cargo_exacto": "Consejera Insular del Area de Accion Social, Inclusion, Voluntariado y Participacion Ciudadana",
        "Nombre_y_apellidos": "Agueda Fumero Roque",
        "Partido_politico": "PP / Grupo Popular",
        "Email_personal_institucional": "aguedafr@tenerife.es",
        "Email_area_consejeria": "accionsocial@tenerife.es",
        "Email_generico_fallback": "atencionciudadana@tenerife.es",
        "Telefono": "922 239 500 / 922 236 870 (FIFEDE empleo)",
        "URL_fuente_principal": "https://www.iass.es/jdownloads/files/portal_de_transparencia/D%20gueda%20Fumero%20Roque%20Grupo%20Popular.pdf",
        "URL_fuente_secundaria": "https://islenior.tenerife.es/aviso-legal/",
        "Nivel_confianza": "Alto",
        "Motivo_confianza": "Ficha oficial de transparencia del IASS confirma cargo, partido y email directo; aviso legal de proyecto del area confirma email del area.",
        "Notas_outreach": "Contacto fuerte para inclusion, discapacidad, sinhogarismo y violencia de genero desde accion social; personalizar por impacto social medible.",
        "Contacto_principal_recomendado": "Si",
        "Comparativa_excel_inicial": "Corregido; el Excel tenia aguedafumero@tenerife.es. La ficha oficial localizada publica aguedafr@tenerife.es.",
    },
    {
        "Isla": "Tenerife",
        "Cabildo": "Cabildo de Tenerife",
        "Area_Consejeria_Servicio": "Empleo, Educacion y Juventud / FIFEDE",
        "Cargo_exacto": "Consejero de Empleo, Educacion y Juventud; FIFEDE como medio propio de empleo/formacion",
        "Nombre_y_apellidos": "Efrain Medina",
        "Partido_politico": "Unknown",
        "Email_personal_institucional": "Unknown",
        "Email_area_consejeria": "informacion@fifede.org",
        "Email_generico_fallback": "atencionciudadana@tenerife.es",
        "Telefono": "922 236 870",
        "URL_fuente_principal": "https://www.tenerife.es/w/el-cabildo-forma-a-mas-de-1.200-personas-desempleadas-a-traves-de-cabildo-emplea-",
        "URL_fuente_secundaria": "https://fifede.org/contacto/",
        "Nivel_confianza": "Medio",
        "Motivo_confianza": "Noticias oficiales identifican al consejero en empleo y formacion; contacto disponible es institucional de FIFEDE, no personal.",
        "Notas_outreach": "Contacto secundario si la propuesta se orienta a formacion online, desempleo y juventud. FIFEDE ya publica formacion sobre IA.",
        "Contacto_principal_recomendado": "No",
        "Comparativa_excel_inicial": "Mejorado; el Excel mezclaba empleo con Accion Social. Se anade canal especifico de FIFEDE.",
    },
    {
        "Isla": "Lanzarote",
        "Cabildo": "Cabildo de Lanzarote",
        "Area_Consejeria_Servicio": "Bienestar Social e Inclusion / Igualdad",
        "Cargo_exacto": "Consejero delegado en Bienestar Social e Inclusion; area de Mayores, Igualdad, Adicciones, Menores, Familia y Sanidad",
        "Nombre_y_apellidos": "Marciano Acuna Betancort",
        "Partido_politico": "Coalicion Canaria",
        "Email_personal_institucional": "Unknown",
        "Email_area_consejeria": "Unknown",
        "Email_generico_fallback": "gobiernoabierto@cabildodelanzarote.com",
        "Telefono": "928 810 100",
        "URL_fuente_principal": "https://transparencia.cabildodelanzarote.com/t/representantes/551",
        "URL_fuente_secundaria": "https://transparencia.cabildodelanzarote.com/storage/uploads/17086785690726.%20Acta%20ordinaria%2029.06.2023.pdf",
        "Nivel_confianza": "Medio",
        "Motivo_confianza": "Portal de transparencia identifica a Marciano Acuna con Bienestar Social e Inclusion; acta oficial delega materias sociales e igualdad, pero no se localizo email especifico.",
        "Notas_outreach": "Recomendado como responsable politico social operativo. Usar telefono central o Gobierno Abierto para pedir traslado a Bienestar Social/Igualdad.",
        "Contacto_principal_recomendado": "Si",
        "Comparativa_excel_inicial": "Corregido/matizado; el Excel asignaba igualdad a Maria Ascension Toledo. La evidencia localizada muestra delegacion social en Marciano Acuna.",
    },
    {
        "Isla": "Lanzarote",
        "Cabildo": "Cabildo de Lanzarote",
        "Area_Consejeria_Servicio": "Educacion, Empleo, Bienestar Social e Inclusion",
        "Cargo_exacto": "Consejera Insular con area de Empleo, Bienestar Social e Inclusion segun organigrama 2023-2027",
        "Nombre_y_apellidos": "Maria Ascension Toledo Hernandez",
        "Partido_politico": "Coalicion Canaria",
        "Email_personal_institucional": "Unknown",
        "Email_area_consejeria": "Unknown",
        "Email_generico_fallback": "gobiernoabierto@cabildodelanzarote.com",
        "Telefono": "928 810 100",
        "URL_fuente_principal": "https://transparencia.cabildodelanzarote.com/storage/uploads/1749043960Legislatura%202023-2027.pdf",
        "URL_fuente_secundaria": "https://transparencia.cabildodelanzarote.com/p/730-1018-organigrama-de-la-entidad-2023-2027",
        "Nivel_confianza": "Medio",
        "Motivo_confianza": "Organigrama oficial 2023-2027 la vincula al area; existen delegaciones posteriores/paralelas que aconsejan contrastar por telefono.",
        "Notas_outreach": "Contacto alternativo si el enfoque se plantea como formacion para empleo o proyecto transversal con Bienestar Social.",
        "Contacto_principal_recomendado": "No",
        "Comparativa_excel_inicial": "Confirmado parcialmente; se mantiene como contacto util, pero no como unica responsable social.",
    },
    {
        "Isla": "Lanzarote",
        "Cabildo": "Cabildo de Lanzarote",
        "Area_Consejeria_Servicio": "Empleo",
        "Cargo_exacto": "Consejero del Grupo de Gobierno con competencias de empleo en organigrama/actas iniciales",
        "Nombre_y_apellidos": "Jesus Alexander Machin Tavio",
        "Partido_politico": "Coalicion Canaria",
        "Email_personal_institucional": "Unknown",
        "Email_area_consejeria": "Unknown",
        "Email_generico_fallback": "gobiernoabierto@cabildodelanzarote.com",
        "Telefono": "928 810 100",
        "URL_fuente_principal": "https://transparencia.cabildodelanzarote.com/t",
        "URL_fuente_secundaria": "https://transparencia.cabildodelanzarote.com/storage/uploads/1690893284Organigrama%20del%20Cabildo%20de%20Lanzarote.pdf",
        "Nivel_confianza": "Bajo",
        "Motivo_confianza": "Persona aparece como grupo de gobierno y organigrama inicial; no se localizo ficha/email ni confirmacion reciente de competencia de empleo tras modificaciones.",
        "Notas_outreach": "Usar solo como contacto de contraste. Para outreach inicial, priorizar Marciano/Maria Ascension y pedir derivacion interna.",
        "Contacto_principal_recomendado": "No",
        "Comparativa_excel_inicial": "No confirmado plenamente; el Excel lo incluia para empleo, pero falta evidencia reciente clara.",
    },
    {
        "Isla": "Fuerteventura",
        "Cabildo": "Cabildo de Fuerteventura",
        "Area_Consejeria_Servicio": "Accion Social, Diversidad, LGTBIQ+, Participacion Ciudadana y Gobierno Abierto",
        "Cargo_exacto": "Consejero insular de Accion Social, Diversidad, LGTBIQ+, Participacion Ciudadana y Gobierno Abierto",
        "Nombre_y_apellidos": "Victor Modesto Alonso Falcon",
        "Partido_politico": "Unknown",
        "Email_personal_institucional": "Unknown",
        "Email_area_consejeria": "familianumerosa@cabildofuer.es",
        "Email_generico_fallback": "sedeelectronica@cabildofuer.es",
        "Telefono": "928 862 300",
        "URL_fuente_principal": "https://www.cabildofuer.es/cabildo/fuerteventura-avanza-en-la-elaboracion-del-i-plan-insular-de-igualdad/",
        "URL_fuente_secundaria": "https://www.cabildofuer.es/cabildo/areas-tematicas/bienestar-social/",
        "Nivel_confianza": "Alto",
        "Motivo_confianza": "Noticias oficiales lo identifican liderando Accion Social e Igualdad; pagina de Bienestar Social aporta contactos tecnicos del area, aunque no email de gabinete.",
        "Notas_outreach": "Contacto principal para igualdad/inclusion. Usar centralita y pedir traslado a Accion Social; email de familia numerosa solo como canal tecnico no ideal.",
        "Contacto_principal_recomendado": "Si",
        "Comparativa_excel_inicial": "Confirmado cargo/persona; se mejora con fuente del Plan Insular de Igualdad y se rebaja calidad del email a tecnico/no politico.",
    },
    {
        "Isla": "Fuerteventura",
        "Cabildo": "Cabildo de Fuerteventura",
        "Area_Consejeria_Servicio": "Empleo, Educacion y Juventud",
        "Cargo_exacto": "Consejera Delegada de Empleo, Educacion y Juventud",
        "Nombre_y_apellidos": "Maria Jesus de la Cruz Montserrat",
        "Partido_politico": "Unknown",
        "Email_personal_institucional": "Unknown",
        "Email_area_consejeria": "empleo@cabildofuer.es / prodae@cabildofuer.es / agenciacolocacion@cabildofuer.es",
        "Email_generico_fallback": "sedeelectronica@cabildofuer.es",
        "Telefono": "928 533 339 / 928 862 484 / 928 862 300",
        "URL_fuente_principal": "https://www.cabildofuer.es/documentos/Empleo/agencia_colocacion_servicios.pdf",
        "URL_fuente_secundaria": "https://sede.cabildofuer.es/portal/transparencia/RecursosWeb/DOCUMENTOS/1/0_1724_1.pdf",
        "Nivel_confianza": "Alto",
        "Motivo_confianza": "Documentos oficiales confirman consejera delegada y canales de la Consejeria de Empleo/Agencia de Colocacion.",
        "Notas_outreach": "Contacto principal si el enfoque se formula como empleabilidad/formacion para mujeres. Priorizar empleo@ o prodae@ antes que agencia de colocacion.",
        "Contacto_principal_recomendado": "Si",
        "Comparativa_excel_inicial": "Confirmado y mejorado; se anaden empleo@cabildofuer.es y prodae@cabildofuer.es como canales mas amplios que solo agencia.",
    },
    {
        "Isla": "La Palma",
        "Cabildo": "Cabildo de La Palma",
        "Area_Consejeria_Servicio": "Accion Social, Igualdad, Diversidad, Vivienda y Salud",
        "Cargo_exacto": "Miembro corporativa con Delegacion Especial en Accion Social, Igualdad, Diversidad, Vivienda y Salud",
        "Nombre_y_apellidos": "Angeles Nieves Fernandez Acosta",
        "Partido_politico": "Unknown",
        "Email_personal_institucional": "Unknown",
        "Email_area_consejeria": "mujer.intervencion@cablapalma.es / mujer.prevencion@cablapalma.es",
        "Email_generico_fallback": "cabildo@cabildodelapalma.es",
        "Telefono": "922 423 100 ext. 4492 / 922 418 000",
        "URL_fuente_principal": "https://transparencia.cabildodelapalma.es/media/r/ayudas-subvenciones/plan-estrategico/2025-2027/Modif%2011%20PES%20Acci%C3%B3n%20Social.pdf",
        "URL_fuente_secundaria": "https://cabildodelapalma.es/es/servicio-de-accion-social",
        "Nivel_confianza": "Alto",
        "Motivo_confianza": "Documento oficial 2025 confirma delegacion exacta; pagina del servicio aporta emails tecnicos de atencion a la mujer.",
        "Notas_outreach": "Contacto principal social/igualdad. En email, solicitar derivacion a la consejera o unidad responsable del programa de igualdad digital.",
        "Contacto_principal_recomendado": "Si",
        "Comparativa_excel_inicial": "Mejorado; el Excel usaba email generico del Cabildo. Se localizan emails especificos de mujer/intervencion y prevencion.",
    },
    {
        "Isla": "La Palma",
        "Cabildo": "Cabildo de La Palma",
        "Area_Consejeria_Servicio": "Formacion y Empleo",
        "Cargo_exacto": "Consejero de Hacienda, Recursos Humanos, Comercio, Formacion, Empleo, Industria y Energia",
        "Nombre_y_apellidos": "Fernando Gonzalez Negrin",
        "Partido_politico": "Unknown",
        "Email_personal_institucional": "Unknown",
        "Email_area_consejeria": "servicio.empleo@cablapalma.es",
        "Email_generico_fallback": "cabildo@cabildodelapalma.es",
        "Telefono": "922 423 100 ext. 4390 / 922 418 000",
        "URL_fuente_principal": "https://cabildodelapalma.es/es/servicios-centrales-de-educacion-empleo-y-formacion",
        "URL_fuente_secundaria": "https://cabildodelapalma.es/es/taxonomy/term/35",
        "Nivel_confianza": "Alto",
        "Motivo_confianza": "Pagina oficial de servicios centrales aporta email de empleo; pagina de Empleo confirma que el area esta dirigida por Fernando Gonzalez.",
        "Notas_outreach": "Contacto principal alternativo para piloto de formacion/empleabilidad con IA.",
        "Contacto_principal_recomendado": "Si",
        "Comparativa_excel_inicial": "Confirmado y mejorado; se sustituye email generico por servicio.empleo@cablapalma.es.",
    },
    {
        "Isla": "La Gomera",
        "Cabildo": "Cabildo Insular de La Gomera",
        "Area_Consejeria_Servicio": "Servicios Sociales, Educacion, Cultura y Deportes / Politica Social e Igualdad",
        "Cargo_exacto": "Vicepresidenta 3a; Consejera Insular del Area de Servicios Sociales, Educacion, Cultura y Deportes",
        "Nombre_y_apellidos": "Rosa Elena Garcia Meneses",
        "Partido_politico": "Agrupacion Socialista Gomera (ASG)",
        "Email_personal_institucional": "rgarcia@lagomera.es",
        "Email_area_consejeria": "Unknown",
        "Email_generico_fallback": "Unknown",
        "Telefono": "922 470 080 / 922 470 081 / 922 470 083 / 922 470 085",
        "URL_fuente_principal": "https://www.lagomera.es/area/presidencia/presidente-y-consejeros-del-cabildo-insular",
        "URL_fuente_secundaria": "https://www.ulpgc.es/sites/default/files/ArchivosULPGC/secretariageneral/Convenios/2026/2026_21_cl_e.pdf",
        "Nivel_confianza": "Alto",
        "Motivo_confianza": "Pagina oficial confirma cargo; convenio publico 2026 confirma email directo de contacto; noticias oficiales la vinculan a politica social e igualdad.",
        "Notas_outreach": "Contacto principal muy adecuado: email directo publico y area social/igualdad. Personalizar con formacion a mujeres, discapacidad y empleo/inclusion.",
        "Contacto_principal_recomendado": "Si",
        "Comparativa_excel_inicial": "Confirmado; se mantiene email directo y se refuerza con convenio publico reciente.",
    },
    {
        "Isla": "La Gomera",
        "Cabildo": "Cabildo Insular de La Gomera",
        "Area_Consejeria_Servicio": "Asuntos Economicos, Obras Publicas, Industria y Movilidad",
        "Cargo_exacto": "Vicepresidenta 2a; Consejera Insular del Area de Asuntos Economicos, Obras Publicas, Industria y Movilidad",
        "Nombre_y_apellidos": "Cristina Ventura Mesa",
        "Partido_politico": "Agrupacion Socialista Gomera (ASG)",
        "Email_personal_institucional": "cventura@lagomera.es",
        "Email_area_consejeria": "Unknown",
        "Email_generico_fallback": "Unknown",
        "Telefono": "922 140 113",
        "URL_fuente_principal": "https://www.lagomera.es/area/presidencia/presidente-y-consejeros-del-cabildo-insular",
        "URL_fuente_secundaria": "https://fr.scribd.com/document/432364793/ORGANIGRAMA",
        "Nivel_confianza": "Bajo",
        "Motivo_confianza": "Cargo actual confirmado, pero el email procede de organigrama antiguo/no oficial replicado en Scribd; empleo no aparece claramente como competencia actual.",
        "Notas_outreach": "No priorizar para SkilLand Mujeres IA salvo enfoque economico/desarrollo. Mantener como posible derivacion interna si Servicios Sociales lo solicita.",
        "Contacto_principal_recomendado": "No",
        "Comparativa_excel_inicial": "Rebajado; el Excel la proponia para empleo, pero no se confirma empleo actual en fuente oficial reciente.",
    },
    {
        "Isla": "El Hierro",
        "Cabildo": "Cabildo de El Hierro",
        "Area_Consejeria_Servicio": "Derechos Sociales, Bienestar Social y Dignidad Personal",
        "Cargo_exacto": "Vicepresidente 2o; Consejero de Derechos Sociales, Bienestar Social y Dignidad Personal",
        "Nombre_y_apellidos": "Amado Carballo Quintero",
        "Partido_politico": "Izquierda Unida-Reunir Canarias (IUC-REUNIR)",
        "Email_personal_institucional": "acarballo@elhierro.es",
        "Email_area_consejeria": "Unknown",
        "Email_generico_fallback": "Unknown",
        "Telefono": "922 550 017 / 922 553 400",
        "URL_fuente_principal": "https://www.elhierro.es/es/amado-carballo-quintero",
        "URL_fuente_secundaria": "https://www.elhierro.es/es/miembros-de-la-corporacion",
        "Nivel_confianza": "Alto",
        "Motivo_confianza": "Ficha oficial y pagina de miembros publican cargo, partido y email directo; marco competencial 2025 confirma area social.",
        "Notas_outreach": "Contacto principal para inclusion, bienestar social, dependencia, mayores y servicios sociales. Muy buen encaje institucional.",
        "Contacto_principal_recomendado": "Si",
        "Comparativa_excel_inicial": "Corregido; el Excel asignaba igualdad/social a Ana Gonzalez. Fuente vigente identifica a Amado Carballo como responsable social.",
    },
    {
        "Isla": "El Hierro",
        "Cabildo": "Cabildo de El Hierro",
        "Area_Consejeria_Servicio": "Empleo y Desarrollo Economico",
        "Cargo_exacto": "Consejera de Empleo y Desarrollo Economico",
        "Nombre_y_apellidos": "Anabel Lopez Garcia",
        "Partido_politico": "Partido Popular (PP)",
        "Email_personal_institucional": "alopez@elhierro.es",
        "Email_area_consejeria": "Unknown",
        "Email_generico_fallback": "Unknown",
        "Telefono": "922 550 017 / 922 553 400",
        "URL_fuente_principal": "https://www.elhierro.es/es/anabel-lopez-garcia",
        "URL_fuente_secundaria": "https://www.elhierro.es/es/miembros-de-la-corporacion",
        "Nivel_confianza": "Alto",
        "Motivo_confianza": "Ficha oficial y pagina de miembros publican cargo, partido y email directo; noticias recientes la confirman en empleo.",
        "Notas_outreach": "Contacto principal para empleabilidad, emprendimiento y formacion con IA orientada a tejido empresarial/desempleo.",
        "Contacto_principal_recomendado": "Si",
        "Comparativa_excel_inicial": "Confirmado; se mantiene email directo y se refuerza con fuente oficial de miembros.",
    },
]


def write_csv() -> None:
    with CSV_PATH.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(ROWS)


def write_xlsx() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Directorio"
    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ROWS:
        ws.append([row[h] for h in HEADERS])
    for col_idx, header in enumerate(HEADERS, start=1):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 55)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    summary = wb.create_sheet("Resumen")
    summary.append(["Metrica", "Valor"])
    summary.append(["Filas totales", len(ROWS)])
    summary.append(["Cabildos cubiertos", 7])
    summary.append(["Contactos principales recomendados", sum(1 for r in ROWS if r["Contacto_principal_recomendado"] == "Si")])
    summary.append(["Confianza alta", sum(1 for r in ROWS if r["Nivel_confianza"] == "Alto")])
    summary.append(["Confianza media", sum(1 for r in ROWS if r["Nivel_confianza"] == "Medio")])
    summary.append(["Confianza baja", sum(1 for r in ROWS if r["Nivel_confianza"] == "Bajo")])
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["B"].width = 18
    wb.save(XLSX_PATH)


def md_table(rows: list[dict[str, str]]) -> str:
    cols = [
        "Isla",
        "Area_Consejeria_Servicio",
        "Nombre_y_apellidos",
        "Cargo_exacto",
        "Email_personal_institucional",
        "Email_area_consejeria",
        "Nivel_confianza",
    ]
    header = "| " + " | ".join(cols) + " |"
    sep = "| " + " | ".join(["---"] * len(cols)) + " |"
    lines = [header, sep]
    for row in rows:
        lines.append("| " + " | ".join(row[c].replace("|", "/") for c in cols) + " |")
    return "\n".join(lines)


def write_markdown() -> None:
    recommended = [r for r in ROWS if r["Contacto_principal_recomendado"] == "Si"]
    by_island: dict[str, list[dict[str, str]]] = {}
    for row in ROWS:
        by_island.setdefault(row["Isla"], []).append(row)

    lines = [
        "# Directorio cabildos igualdad social",
        "",
        "## Resumen ejecutivo",
        "",
        "Primera investigacion publica para localizar responsables utiles de los 7 Cabildos de Canarias en igualdad, accion social, inclusion, bienestar social, empleo y areas conectadas. Se han priorizado cargos y canales institucionales confirmados en portales oficiales, organigramas, paginas de area, documentos de transparencia y noticias institucionales recientes.",
        "",
        f"- Cabildos cubiertos: 7/7.",
        f"- Contactos totales incluidos: {len(ROWS)}.",
        f"- Contactos principales recomendados: {len(recommended)}.",
        f"- Contactos con confianza alta: {sum(1 for r in ROWS if r['Nivel_confianza'] == 'Alto')}.",
        "- Mejoras clave frente al Excel inicial: correccion de El Hierro social hacia Amado Carballo; sustitucion de Presidencia por Direccion Insular de Igualdad en Tenerife; emails especificos localizados para La Palma, Fuerteventura y Tenerife; matizacion de Lanzarote por delegaciones cruzadas.",
        "",
        "## Tabla final de contactos recomendados",
        "",
        md_table(recommended),
        "",
        "## Detalle por Cabildo",
        "",
    ]

    for island, rows in by_island.items():
        lines.extend([f"### {island}", ""])
        for row in rows:
            email = row["Email_personal_institucional"]
            if email == "Unknown":
                email = row["Email_area_consejeria"]
            if email == "Unknown":
                email = row["Email_generico_fallback"]
            lines.extend(
                [
                    f"**{row['Nombre_y_apellidos']}**",
                    f"- Cargo: {row['Cargo_exacto']}",
                    f"- Area: {row['Area_Consejeria_Servicio']}",
                    f"- Email recomendado: {email}",
                    f"- Telefono: {row['Telefono']}",
                    f"- Fuente principal: {row['URL_fuente_principal']}",
                    f"- Fuente secundaria: {row['URL_fuente_secundaria']}",
                    f"- Confianza: {row['Nivel_confianza']} - {row['Motivo_confianza']}",
                    f"- Dudas pendientes: {'Ninguna critica' if row['Nivel_confianza'] == 'Alto' else 'Conviene validar por telefono o portal de contacto antes de outreach masivo.'}",
                    f"- Notas outreach: {row['Notas_outreach']}",
                    "",
                ]
            )

    lines.extend(
        [
            "## Comparativa rapida contra el Excel inicial",
            "",
            "### Datos confirmados",
            "",
            "- Gran Canaria: Isabel Mena Alonso en Politica Social/Igualdad y Juan Diaz Sanchez en Empleo/Desarrollo Local.",
            "- Fuerteventura: Victor Alonso en Accion Social/Igualdad y Maria Jesus de la Cruz Montserrat en Empleo/Educacion/Juventud.",
            "- La Palma: Angeles Nieves Fernandez Acosta en Accion Social/Igualdad y Fernando Gonzalez Negrin en Formacion/Empleo.",
            "- La Gomera: Rosa Elena Garcia Meneses como responsable social/igualdad con email directo rgarcia@lagomera.es.",
            "- El Hierro: Anabel Lopez Garcia como Empleo y Desarrollo Economico.",
            "",
            "### Datos corregidos",
            "",
            "- Tenerife: para igualdad se recomienda Patricia Leon Perez/Direccion Insular de Igualdad y Diversidad en lugar de Presidencia como primer contacto operativo.",
            "- Tenerife: email directo de Agueda Fumero corregido a aguedafr@tenerife.es segun ficha oficial del IASS.",
            "- Lanzarote: se matiza el responsable social. Marciano Acuna Betancort aparece como consejero delegado de Bienestar Social e Inclusion; Maria Ascension Toledo sigue siendo contacto transversal util por organigrama.",
            "- El Hierro: el responsable social vigente es Amado Carballo Quintero, no Ana Gonzalez Gonzalez.",
            "",
            "### Datos no encontrados",
            "",
            "- Emails personales directos para Isabel Mena, Juan Diaz, Patricia Leon, Efrain Medina, Marciano Acuna, Maria Ascension Toledo, Jesus Alexander Machin, Victor Alonso, Maria Jesus de la Cruz, Angeles Fernandez y Fernando Gonzalez.",
            "- Email especifico publico de Bienestar Social/Igualdad en Lanzarote.",
            "- Confirmacion oficial reciente de competencia de empleo de Jesus Alexander Machin tras modificaciones de organigrama en Lanzarote.",
            "- Partido politico publico en fuentes oficiales para varios cargos fuera de paginas de miembros con partido visible.",
            "",
            "### Mejoras conseguidas",
            "",
            "- Se han sustituido varios correos genericos por emails de area o servicio: igualdadydiversidad@tenerife.es, accionsocial@tenerife.es, informacion@fifede.org, empleo@cabildofuer.es, prodae@cabildofuer.es, mujer.intervencion@cablapalma.es, mujer.prevencion@cablapalma.es, servicio.empleo@cablapalma.es.",
            "- Se han incorporado fuentes secundarias para justificar cargos y contactos.",
            "- Se han marcado contactos principales por Cabildo y contactos alternativos para empleo/formacion.",
            "- Se han anotado gaps accionables para validacion telefonica previa al envio del dossier.",
        ]
    )
    MD_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_notes() -> None:
    lines = [
        "# Scraping notes",
        "",
        "Fecha de investigacion: 2026-05-27.",
        "",
        "## Metodo",
        "",
        "- Lectura del contexto estrategico SkilLand IA Mujeres y del Excel base.",
        "- Busqueda publica en portales oficiales de Cabildos, portales de transparencia, paginas de area, documentos PDF institucionales y noticias oficiales.",
        "- Priorizacion de email segun jerarquia: directo personal, area/consejeria/servicio, gabinete/secretaria, generico.",
        "- No se inventaron emails ni se dedujeron patrones.",
        "",
        "## Fuentes principales consultadas",
        "",
    ]
    for row in ROWS:
        lines.append(f"- {row['Isla']} / {row['Nombre_y_apellidos']}: {row['URL_fuente_principal']} | {row['URL_fuente_secundaria']}")
    lines.extend(
        [
            "",
            "## Incidencias y decisiones",
            "",
            "- Tenerife: la web de miembros electos de transparencia no expuso contenido util en la vista textual; se usaron paginas oficiales de area, IASS y noticias oficiales.",
            "- Lanzarote: hay aparente solapamiento entre organigrama y delegaciones. Se incluye a Marciano Acuna como principal social por ficha de representante/delegacion y a Maria Ascension Toledo como contacto transversal alternativo.",
            "- Fuerteventura: no se localizo email politico directo de Victor Alonso ni Maria Jesus de la Cruz. Se priorizaron emails de servicios oficiales de empleo y canales tecnicos/sociales.",
            "- La Gomera: el email cventura@lagomera.es solo se encontro en organigrama antiguo/no oficial replicado; se mantiene con confianza baja y no como contacto recomendado.",
            "- El Hierro: fuente oficial 2025-2026 desplaza el contacto social del Excel desde Ana Gonzalez hacia Amado Carballo.",
            "",
            "## Gaps pendientes",
            "",
            "- Validar por telefono o formulario interno los emails de area que no son gabinete directo.",
            "- Pedir derivacion explicita a gabinete/consejeria en Lanzarote y Fuerteventura antes de hacer envios comerciales.",
            "- Si el funnel requiere maxima precision, hacer una segunda pasada manual por registros de decretos de delegacion y directorios internos descargables.",
        ]
    )
    NOTES_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    write_csv()
    write_xlsx()
    write_markdown()
    write_notes()
    print(CSV_PATH)
    print(XLSX_PATH)
    print(MD_PATH)
    print(NOTES_PATH)


if __name__ == "__main__":
    main()
