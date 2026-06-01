from __future__ import annotations

import csv
import json
import re
import sys
import time
import urllib.parse
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

import lxml.html
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUT_DIR = Path("04_outputs/skilland-ia-mujeres/ayuntamientos_contactos_igualdad_social")
CHECKPOINTS_DIR = OUT_DIR / "checkpoints"
SCRATCH_DIR = Path("05_scratch/ayuntamientos_contactos_igualdad_social")
SEED_CSV = OUT_DIR / "municipios_canarias_seed.csv"
DIRECTORY_CSV = OUT_DIR / "directorio_ayuntamientos_igualdad_social.csv"
DIRECTORY_XLSX = OUT_DIR / "directorio_ayuntamientos_igualdad_social.xlsx"
DIRECTORY_MD = OUT_DIR / "directorio_ayuntamientos_igualdad_social.md"
SOURCES_JSON = OUT_DIR / "fuentes_ayuntamientos.json"
COVERAGE_MD = OUT_DIR / "coverage_report.md"
NOTES_MD = OUT_DIR / "scraping_notes_ayuntamientos.md"
RAW_CONTACTS_CSV = SCRATCH_DIR / "directorio_ayuntamientos_igualdad_social_raw.csv"
CRAWL_JSON = SCRATCH_DIR / "crawl_candidates.json"
CRAWL_MD = SCRATCH_DIR / "crawl_candidates.md"

SEED_HEADERS = [
    "Isla",
    "Municipio",
    "Ayuntamiento",
    "URL oficial ayuntamiento",
    "URL transparencia",
    "Estado busqueda",
]

DIRECTORY_HEADERS = [
    "Isla",
    "Municipio",
    "Ayuntamiento",
    "Area / Concejalia / Servicio",
    "Cargo exacto",
    "Nombre y apellidos",
    "Partido politico",
    "Email institucional personal",
    "Email del area / concejalia / servicio",
    "Email generico municipal",
    "Telefono directo",
    "Telefono general",
    "URL fuente principal",
    "URL fuente secundaria",
    "Tipo de fuente principal",
    "Nivel de confianza",
    "Motivo del nivel de confianza",
    "Contacto recomendado",
    "Prioridad outreach",
    "Notas utiles para outreach",
    "Dudas pendientes / validacion manual necesaria",
]

KEYWORDS = [
    "igualdad",
    "mujer",
    "social",
    "servicios-sociales",
    "servicios_sociales",
    "bienestar",
    "accion-social",
    "accion_social",
    "inclusion",
    "empleo",
    "desarrollo-local",
    "desarrollo_local",
    "formacion",
    "juventud",
    "mayores",
    "dependencia",
    "diversidad",
    "discapacidad",
    "violencia",
    "concejal",
    "concejalia",
    "corporacion",
    "gobierno",
    "equipo-de-gobierno",
    "organigrama",
    "contacto",
    "transparencia",
]

PATH_BOOST = {
    "igualdad": 120,
    "mujer": 110,
    "servicios-sociales": 100,
    "social": 90,
    "bienestar": 90,
    "inclusion": 85,
    "empleo": 80,
    "desarrollo-local": 70,
    "formacion": 65,
    "juventud": 50,
    "corporacion": 45,
    "gobierno": 40,
    "concejal": 40,
    "contacto": 25,
    "transparencia": 20,
}

ISLAND_ORDER = [
    "El Hierro",
    "La Gomera",
    "La Palma",
    "Fuerteventura",
    "Lanzarote",
    "Gran Canaria",
    "Tenerife",
]

SEED_ROWS = [
    ("El Hierro", "El Pinar de El Hierro", "Ayuntamiento de El Pinar de El Hierro", "http://www.elpinardeelhierro.com", "Unknown", "pendiente"),
    ("El Hierro", "Frontera", "Ayuntamiento de La Frontera", "http://www.aytofrontera.org/", "Unknown", "pendiente"),
    ("El Hierro", "Valverde", "Ayuntamiento de Valverde", "http://www.aytovalverde.org/", "Unknown", "pendiente"),
    ("La Gomera", "Agulo", "Ayuntamiento de Agulo", "http://www.agulo.org/", "Unknown", "pendiente"),
    ("La Gomera", "Alajero", "Ayuntamiento de Alajero", "http://www.ayuntamientoalajero.es/", "Unknown", "pendiente"),
    ("La Gomera", "Hermigua", "Ayuntamiento de Hermigua", "http://www.hermigua.es/", "Unknown", "pendiente"),
    ("La Gomera", "San Sebastian de La Gomera", "Ayuntamiento de San Sebastian de La Gomera", "http://www.sansebastiangomera.org/", "Unknown", "pendiente"),
    ("La Gomera", "Valle Gran Rey", "Ayuntamiento de Valle Gran Rey", "http://www.vallegranrey.es/", "Unknown", "pendiente"),
    ("La Gomera", "Vallehermoso", "Ayuntamiento de Vallehermoso", "http://www.ayuntamientovallehermoso.org/", "Unknown", "pendiente"),
    ("La Palma", "Barlovento", "Ayuntamiento de Barlovento", "http://www.barlovento.es/", "Unknown", "pendiente"),
    ("La Palma", "Brena Alta", "Ayuntamiento de Brena Alta", "http://www.balta.org/", "Unknown", "pendiente"),
    ("La Palma", "Brena Baja", "Ayuntamiento de Brena Baja", "http://www.bbaja.es/", "Unknown", "pendiente"),
    ("La Palma", "El Paso", "Ayuntamiento de El Paso", "http://www.elpaso.es/", "Unknown", "pendiente"),
    ("La Palma", "Fuencaliente", "Ayuntamiento de Fuencaliente de La Palma", "http://www.fuencalientedelapalma.es/", "Unknown", "pendiente"),
    ("La Palma", "Garafia", "Ayuntamiento de Garafia", "http://www.garafia.es/", "Unknown", "pendiente"),
    ("La Palma", "Los Llanos de Aridane", "Ayuntamiento de Los Llanos de Aridane", "http://www.aridane.org/", "Unknown", "pendiente"),
    ("La Palma", "Puntagorda", "Ayuntamiento de Puntagorda", "http://www.puntagorda.es/", "Unknown", "pendiente"),
    ("La Palma", "Puntallana", "Ayuntamiento de Puntallana", "http://www.puntallana.es/", "Unknown", "pendiente"),
    ("La Palma", "San Andres y Sauces", "Ayuntamiento de San Andres y Sauces", "http://www.sanandresysauces.es", "Unknown", "pendiente"),
    ("La Palma", "Santa Cruz de La Palma", "Ayuntamiento de Santa Cruz de La Palma", "http://www.santacruzdelapalma.es/", "Unknown", "pendiente"),
    ("La Palma", "Tazacorte", "Ayuntamiento de Tazacorte", "http://www.tazacorte.es/", "Unknown", "pendiente"),
    ("La Palma", "Tijarafe", "Ayuntamiento de Tijarafe", "http://www.tijarafe.es", "Unknown", "pendiente"),
    ("La Palma", "Villa de Mazo", "Ayuntamiento de Villa de Mazo", "http://www.villademazo.com", "Unknown", "pendiente"),
    ("Fuerteventura", "Antigua", "Ayuntamiento de Antigua", "http://www.ayto-antigua.es", "Unknown", "pendiente"),
    ("Fuerteventura", "Betancuria", "Ayuntamiento de Betancuria", "http://www.aytobetancuria.org/", "Unknown", "pendiente"),
    ("Fuerteventura", "La Oliva", "Ayuntamiento de La Oliva", "http://www.laoliva.es/", "Unknown", "pendiente"),
    ("Fuerteventura", "Pajara", "Ayuntamiento de Pajara", "http://www.pajara.es", "Unknown", "pendiente"),
    ("Fuerteventura", "Puerto del Rosario", "Ayuntamiento de Puerto del Rosario", "http://www.puertodelrosario.org", "Unknown", "pendiente"),
    ("Fuerteventura", "Tuineje", "Ayuntamiento de Tuineje", "http://www.ayuntamientotuineje.com", "Unknown", "pendiente"),
    ("Lanzarote", "Arrecife", "Ayuntamiento de Arrecife", "http://www.arrecife.es", "Unknown", "pendiente"),
    ("Lanzarote", "Haria", "Ayuntamiento de Haria", "http://www.ayuntamientodeharia.com", "Unknown", "pendiente"),
    ("Lanzarote", "San Bartolome", "Ayuntamiento de San Bartolome", "http://www.sanbartolome.es", "Unknown", "pendiente"),
    ("Lanzarote", "Teguise", "Ayuntamiento de Teguise", "http://www.teguise.es", "Unknown", "pendiente"),
    ("Lanzarote", "Tias", "Ayuntamiento de Tias", "http://www.ayuntamientodetias.es", "Unknown", "pendiente"),
    ("Lanzarote", "Tinajo", "Ayuntamiento de Tinajo", "http://www.tinajo.es", "Unknown", "pendiente"),
    ("Lanzarote", "Yaiza", "Ayuntamiento de Yaiza", "http://www.yaiza.org", "Unknown", "pendiente"),
    ("Gran Canaria", "Agaete", "Ayuntamiento de Agaete", "http://www.aytoagaete.es", "Unknown", "pendiente"),
    ("Gran Canaria", "Aguimes", "Ayuntamiento de Aguimes", "http://www.aguimes.es/", "Unknown", "pendiente"),
    ("Gran Canaria", "Artenara", "Ayuntamiento de Artenara", "http://www.artenara.es/", "Unknown", "pendiente"),
    ("Gran Canaria", "Arucas", "Ayuntamiento de Arucas", "http://www.arucas.org/", "Unknown", "pendiente"),
    ("Gran Canaria", "Firgas", "Ayuntamiento de Firgas", "http://www.firgas.es/", "Unknown", "pendiente"),
    ("Gran Canaria", "Galdar", "Ayuntamiento de Galdar", "http://www.galdar.es/", "Unknown", "pendiente"),
    ("Gran Canaria", "Ingenio", "Ayuntamiento de Ingenio", "http://www.ingenio.es/", "Unknown", "pendiente"),
    ("Gran Canaria", "La Aldea de San Nicolas", "Ayuntamiento de La Aldea de San Nicolas", "http://www.laaldeadesannicolas.es/", "Unknown", "pendiente"),
    ("Gran Canaria", "Las Palmas de Gran Canaria", "Ayuntamiento de Las Palmas de Gran Canaria", "http://www.laspalmasgc.es/", "Unknown", "pendiente"),
    ("Gran Canaria", "Mogan", "Ayuntamiento de Mogan", "http://www.mogan.es/", "Unknown", "pendiente"),
    ("Gran Canaria", "Moya", "Ayuntamiento de Villa de Moya", "http://www.villademoya.es/", "Unknown", "pendiente"),
    ("Gran Canaria", "San Bartolome de Tirajana", "Ayuntamiento de San Bartolome de Tirajana", "http://www.maspalomas.com/", "Unknown", "pendiente"),
    ("Gran Canaria", "Santa Brigida", "Ayuntamiento de Santa Brigida", "http://www.santabrigida.es/", "Unknown", "pendiente"),
    ("Gran Canaria", "Santa Lucia de Tirajana", "Ayuntamiento de Santa Lucia de Tirajana", "http://www.santaluciagc.com", "Unknown", "pendiente"),
    ("Gran Canaria", "Santa Maria de Guia", "Ayuntamiento de Santa Maria de Guia", "http://www.santamariadeguia.es/", "Unknown", "pendiente"),
    ("Gran Canaria", "Tejeda", "Ayuntamiento de Tejeda", "http://www.tejeda.es/", "Unknown", "pendiente"),
    ("Gran Canaria", "Telde", "Ayuntamiento de Telde", "http://www.telde.es/", "Unknown", "pendiente"),
    ("Gran Canaria", "Teror", "Ayuntamiento de Teror", "http://www.teror.es/", "Unknown", "pendiente"),
    ("Gran Canaria", "Valleseco", "Ayuntamiento de Valleseco", "http://www.valleseco.es/", "Unknown", "pendiente"),
    ("Gran Canaria", "Valsequillo", "Ayuntamiento de Valsequillo de Gran Canaria", "http://www.valsequillogc.es/", "Unknown", "pendiente"),
    ("Gran Canaria", "Vega de San Mateo", "Ayuntamiento de Vega de San Mateo", "http://www.vegasanmateo.es/", "Unknown", "pendiente"),
    ("Tenerife", "Adeje", "Ayuntamiento de Adeje", "http://www.adeje.es/", "Unknown", "pendiente"),
    ("Tenerife", "Arafo", "Ayuntamiento de Arafo", "http://www.arafo.es/", "Unknown", "pendiente"),
    ("Tenerife", "Arico", "Ayuntamiento de Arico", "http://www.ayuntamientodearico.com", "Unknown", "pendiente"),
    ("Tenerife", "Arona", "Ayuntamiento de Arona", "http://www.arona.org/", "Unknown", "pendiente"),
    ("Tenerife", "Buenavista del Norte", "Ayuntamiento de Buenavista del Norte", "http://www.buenavistadelnorte.com/", "Unknown", "pendiente"),
    ("Tenerife", "Candelaria", "Ayuntamiento de Candelaria", "http://www.candelaria.es/", "Unknown", "pendiente"),
    ("Tenerife", "El Rosario", "Ayuntamiento de El Rosario", "http://www.ayuntamientoelrosario.org/", "Unknown", "pendiente"),
    ("Tenerife", "El Sauzal", "Ayuntamiento de El Sauzal", "http://www.elsauzal.es/", "Unknown", "pendiente"),
    ("Tenerife", "El Tanque", "Ayuntamiento de El Tanque", "http://www.eltanque.es", "Unknown", "pendiente"),
    ("Tenerife", "Fasnia", "Ayuntamiento de Fasnia", "http://www.ayuntamientodefasnia.es/", "Unknown", "pendiente"),
    ("Tenerife", "Garachico", "Ayuntamiento de Garachico", "http://www.garachico.es/", "Unknown", "pendiente"),
    ("Tenerife", "Granadilla de Abona", "Ayuntamiento de Granadilla de Abona", "http://www.granadilladeabona.org", "Unknown", "pendiente"),
    ("Tenerife", "Guia de Isora", "Ayuntamiento de Guia de Isora", "http://www.guiadeisora.org/", "Unknown", "pendiente"),
    ("Tenerife", "Guimar", "Ayuntamiento de Guimar", "http://www.guimar.es/", "Unknown", "pendiente"),
    ("Tenerife", "Icod de los Vinos", "Ayuntamiento de Icod de los Vinos", "http://www.icoddelosvinos.es/", "Unknown", "pendiente"),
    ("Tenerife", "La Guancha", "Ayuntamiento de La Guancha", "http://www.laguancha.es", "Unknown", "pendiente"),
    ("Tenerife", "La Laguna", "Ayuntamiento de San Cristobal de La Laguna", "http://www.aytolalaguna.org/", "Unknown", "pendiente"),
    ("Tenerife", "La Matanza de Acentejo", "Ayuntamiento de La Matanza de Acentejo", "http://www.matanceros.com", "Unknown", "pendiente"),
    ("Tenerife", "La Orotava", "Ayuntamiento de La Orotava", "http://www.villadelaorotava.org/", "Unknown", "pendiente"),
    ("Tenerife", "La Victoria de Acentejo", "Ayuntamiento de La Victoria de Acentejo", "http://www.lavictoriadeacentejo.es", "Unknown", "pendiente"),
    ("Tenerife", "Los Realejos", "Ayuntamiento de Los Realejos", "http://www.ayto-realejos.es/", "Unknown", "pendiente"),
    ("Tenerife", "Los Silos", "Ayuntamiento de Los Silos", "http://www.lossilos.es", "Unknown", "pendiente"),
    ("Tenerife", "Puerto de la Cruz", "Ayuntamiento de Puerto de la Cruz", "http://www.puertodelacruz.es/", "Unknown", "pendiente"),
    ("Tenerife", "San Juan de la Rambla", "Ayuntamiento de San Juan de la Rambla", "http://www.aytosanjuandelarambla.es/", "Unknown", "pendiente"),
    ("Tenerife", "San Miguel de Abona", "Ayuntamiento de San Miguel de Abona", "http://www.sanmigueldeabona.es/", "Unknown", "pendiente"),
    ("Tenerife", "Santa Cruz de Tenerife", "Ayuntamiento de Santa Cruz de Tenerife", "http://www.santacruzdetenerife.es/", "Unknown", "pendiente"),
    ("Tenerife", "Santa Ursula", "Ayuntamiento de Santa Ursula", "http://www.santaursula.es", "Unknown", "pendiente"),
    ("Tenerife", "Santiago del Teide", "Ayuntamiento de Santiago del Teide", "http://www.santiagodelteide.org/", "Unknown", "pendiente"),
    ("Tenerife", "Tacoronte", "Ayuntamiento de Tacoronte", "http://www.tacoronte.es/", "Unknown", "pendiente"),
    ("Tenerife", "Tegueste", "Ayuntamiento de Tegueste", "http://www.tegueste.org/", "Unknown", "pendiente"),
    ("Tenerife", "Vilaflor de Chasna", "Ayuntamiento de Vilaflor de Chasna", "http://www.vilaflordechasna.es/", "Unknown", "pendiente"),
]


EMAIL_RE = re.compile(r"(?i)([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})")
PHONE_RE = re.compile(r"(?:(?:\+34|0034)\s*)?(?:\d[\s.-]?){9,}")


@dataclass
class CandidatePage:
    island: str
    municipality: str
    official_url: str
    url: str
    score: int
    source: str
    title: str
    headings: list[str]
    emails: list[str]
    phones: list[str]
    keyword_hits: list[str]
    snippet: str


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def slug(text: str) -> str:
    text = text.lower().strip()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ü": "u",
        "ñ": "n",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def ensure_dirs() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    CHECKPOINTS_DIR.mkdir(parents=True, exist_ok=True)
    SCRATCH_DIR.mkdir(parents=True, exist_ok=True)


def build_seed_rows() -> list[dict[str, str]]:
    rows = []
    for island, municipality, ayuntamiento, official_url, transparencia_url, state in SEED_ROWS:
        rows.append(
            {
                "Isla": island,
                "Municipio": municipality,
                "Ayuntamiento": ayuntamiento,
                "URL oficial ayuntamiento": official_url,
                "URL transparencia": transparencia_url,
                "Estado busqueda": state,
            }
        )
    return rows


def write_csv(path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def fetch(session: requests.Session, url: str, timeout: int = 25) -> str | None:
    try:
        response = session.get(url, timeout=timeout, headers={"User-Agent": "Mozilla/5.0 municipal-directory-research"})
        response.raise_for_status()
        response.encoding = response.apparent_encoding or response.encoding
        if "text/html" in response.headers.get("content-type", "") or response.text.startswith("<"):
            return response.text
        return None
    except Exception:
        return None


def canonicalize(base_url: str, maybe_url: str) -> str | None:
    joined = urllib.parse.urljoin(base_url, maybe_url)
    parsed = urllib.parse.urlparse(joined)
    if parsed.scheme not in {"http", "https"}:
        return None
    return urllib.parse.urlunparse((parsed.scheme, parsed.netloc, parsed.path or "/", parsed.params, parsed.query, ""))


def same_domain(left: str, right: str) -> bool:
    return urllib.parse.urlparse(left).netloc.lower().replace("www.", "") == urllib.parse.urlparse(right).netloc.lower().replace("www.", "")


def extract_links(base_url: str, html_text: str) -> list[str]:
    links: list[str] = []
    try:
        doc = lxml.html.fromstring(html_text)
    except Exception:
        return links
    for node in doc.xpath("//a[@href]"):
        href = node.get("href") or ""
        url = canonicalize(base_url, href)
        if not url:
            continue
        if same_domain(base_url, url):
            links.append(url)
    return links


def extract_sitemaps(base_url: str, robots_text: str | None) -> list[str]:
    sitemaps: list[str] = []
    if robots_text:
        for line in robots_text.splitlines():
            if line.lower().startswith("sitemap:"):
                maybe = line.split(":", 1)[1].strip()
                url = canonicalize(base_url, maybe)
                if url:
                    sitemaps.append(url)
    fallback_paths = [
        "/sitemap.xml",
        "/sitemap_index.xml",
        "/wp-sitemap.xml",
        "/sitemap-index.xml",
    ]
    for path in fallback_paths:
        sitemaps.append(canonicalize(base_url, path))
    return [url for url in sitemaps if url]


def parse_sitemap_xml(xml_text: str) -> list[str]:
    try:
        root = ET.fromstring(xml_text.encode("utf-8"))
    except Exception:
        return []
    urls: list[str] = []
    for elem in root.iter():
        if elem.tag.endswith("loc") and elem.text:
            urls.append(elem.text.strip())
    return urls


def keyword_score(url: str, text: str) -> tuple[int, list[str]]:
    score = 0
    hits: list[str] = []
    lowered_url = url.lower()
    lowered_text = text.lower()
    for keyword, boost in PATH_BOOST.items():
        if keyword in lowered_url:
            score += boost
            hits.append(keyword)
    for keyword in KEYWORDS:
        if keyword in lowered_text and keyword not in hits:
            score += 15
            hits.append(keyword)
    if "@" in lowered_text:
        score += 10
    return score, hits


def extract_page_metadata(url: str, html_text: str, island: str, municipality: str, official_url: str, source: str) -> CandidatePage:
    title = ""
    headings: list[str] = []
    text = ""
    try:
        doc = lxml.html.fromstring(html_text)
        title = normalize_text(" ".join(doc.xpath("//title/text()")))
        headings = [normalize_text(value) for value in doc.xpath("//h1/text() | //h2/text() | //h3/text()")]
        text = normalize_text(" ".join(doc.xpath("//body//text()")))
    except Exception:
        text = normalize_text(html_text)
    emails = sorted(set(match.lower() for match in EMAIL_RE.findall(html_text)))
    phones = []
    for match in PHONE_RE.findall(text):
        cleaned = normalize_text(match)
        digits = re.sub(r"\D", "", cleaned)
        if len(digits) >= 9:
            phones.append(cleaned)
    phones = sorted(set(phones))
    score, hits = keyword_score(url, " ".join([title, *headings, text[:4000]]))
    snippet_parts = []
    for keyword in ["igualdad", "mujer", "servicios sociales", "bienestar", "empleo", "desarrollo local", "concejal", "corporacion"]:
        pos = text.lower().find(keyword)
        if pos != -1:
            start = max(pos - 100, 0)
            end = min(pos + 240, len(text))
            snippet_parts.append(text[start:end])
            if len(snippet_parts) == 2:
                break
    snippet = normalize_text(" ... ".join(snippet_parts))[:700]
    return CandidatePage(
        island=island,
        municipality=municipality,
        official_url=official_url,
        url=url,
        score=score,
        source=source,
        title=title,
        headings=headings[:8],
        emails=emails[:10],
        phones=phones[:10],
        keyword_hits=hits[:10],
        snippet=snippet,
    )


def crawl_candidates(limit_islands: set[str] | None = None, limit_municipalities: set[str] | None = None) -> dict[str, list[dict[str, object]]]:
    session = requests.Session()
    results: dict[str, list[dict[str, object]]] = {}
    seed_rows = build_seed_rows()
    for row in seed_rows:
        island = row["Isla"]
        municipality = row["Municipio"]
        if limit_islands and island not in limit_islands:
            continue
        if limit_municipalities and municipality not in limit_municipalities:
            continue
        official_url = row["URL oficial ayuntamiento"]
        robots_text = fetch(session, urllib.parse.urljoin(official_url, "/robots.txt"), timeout=15)
        candidate_urls: set[str] = {official_url}
        for sitemap_url in extract_sitemaps(official_url, robots_text):
            sitemap_text = fetch(session, sitemap_url, timeout=15)
            if not sitemap_text:
                continue
            for site_url in parse_sitemap_xml(sitemap_text):
                if same_domain(official_url, site_url):
                    candidate_urls.add(site_url)
        home_html = fetch(session, official_url, timeout=20)
        if home_html:
            for link in extract_links(official_url, home_html):
                candidate_urls.add(link)
        ranked: list[tuple[int, str, str]] = []
        for candidate_url in candidate_urls:
            lowered = candidate_url.lower()
            score = 0
            source = "link"
            for keyword, boost in PATH_BOOST.items():
                if keyword in lowered:
                    score += boost
            if score == 0 and candidate_url == official_url:
                score = 5
            if "sitemap" in lowered:
                continue
            if score > 0:
                if candidate_url not in extract_links(official_url, home_html or ""):
                    source = "sitemap"
                ranked.append((score, candidate_url, source))
        ranked.sort(reverse=True)
        pages: list[dict[str, object]] = []
        seen_urls: set[str] = set()
        for _, candidate_url, source in ranked[:12]:
            if candidate_url in seen_urls:
                continue
            seen_urls.add(candidate_url)
            html_text = fetch(session, candidate_url, timeout=20)
            if not html_text:
                continue
            page = extract_page_metadata(candidate_url, html_text, island, municipality, official_url, source)
            if page.score <= 0:
                continue
            pages.append(
                {
                    "url": page.url,
                    "score": page.score,
                    "source": page.source,
                    "title": page.title,
                    "headings": page.headings,
                    "emails": page.emails,
                    "phones": page.phones,
                    "keyword_hits": page.keyword_hits,
                    "snippet": page.snippet,
                }
            )
            time.sleep(0.2)
        pages.sort(key=lambda item: int(item["score"]), reverse=True)
        results[f"{island}::{municipality}"] = pages[:12]
    return results


def write_crawl_reports(candidates: dict[str, list[dict[str, object]]]) -> None:
    with CRAWL_JSON.open("w", encoding="utf-8") as handle:
        json.dump(candidates, handle, ensure_ascii=False, indent=2)
    lines = ["# Crawl candidates", "", "Resumen de paginas candidatas por municipio.", ""]
    for key, pages in candidates.items():
        island, municipality = key.split("::", 1)
        lines.append(f"## {island} / {municipality}")
        if not pages:
            lines.append("")
            lines.append("- Sin candidatos con keywords detectadas.")
            lines.append("")
            continue
        lines.append("")
        for page in pages[:5]:
            lines.append(f"- Score {page['score']} | {page['url']}")
            title = page.get("title") or "Unknown"
            lines.append(f"  Title: {title}")
            emails = ", ".join(page.get("emails") or []) or "Unknown"
            phones = ", ".join(page.get("phones") or []) or "Unknown"
            lines.append(f"  Emails: {emails}")
            lines.append(f"  Phones: {phones}")
            if page.get("headings"):
                lines.append(f"  Headings: {' | '.join(page['headings'][:4])}")
            if page.get("snippet"):
                lines.append(f"  Snippet: {page['snippet']}")
        lines.append("")
    CRAWL_MD.write_text("\n".join(lines), encoding="utf-8")


def build_sources_index(rows: list[dict[str, str]]) -> dict[str, dict[str, object]]:
    grouped: dict[str, dict[str, object]] = {}
    for row in rows:
        key = f"{row['Isla']}::{row['Municipio']}"
        bucket = grouped.setdefault(
            key,
            {
                "isla": row["Isla"],
                "municipio": row["Municipio"],
                "ayuntamiento": row["Ayuntamiento"],
                "fuentes": [],
            },
        )
        for field in ["URL fuente principal", "URL fuente secundaria"]:
            url = row.get(field, "").strip()
            if url and url != "Unknown":
                bucket["fuentes"].append(
                    {
                        "url": url,
                        "tipo": row.get("Tipo de fuente principal", "Unknown"),
                        "area": row.get("Area / Concejalia / Servicio", "Unknown"),
                        "cargo": row.get("Cargo exacto", "Unknown"),
                    }
                )
    for payload in grouped.values():
        unique: list[dict[str, object]] = []
        seen: set[str] = set()
        for item in payload["fuentes"]:
            if item["url"] in seen:
                continue
            seen.add(item["url"])
            unique.append(item)
        payload["fuentes"] = unique
    return grouped


def write_xlsx(rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ayuntamientos"
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(DIRECTORY_HEADERS, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(DIRECTORY_HEADERS, start=1):
            value = row.get(header, "Unknown")
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        "A": 15,
        "B": 24,
        "C": 32,
        "D": 34,
        "E": 32,
        "F": 28,
        "G": 22,
        "H": 28,
        "I": 34,
        "J": 28,
        "K": 18,
        "L": 18,
        "M": 42,
        "N": 42,
        "O": 20,
        "P": 14,
        "Q": 42,
        "R": 16,
        "S": 18,
        "T": 40,
        "U": 36,
    }
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A2"
    workbook.save(DIRECTORY_XLSX)


def write_markdown(rows: list[dict[str, str]]) -> None:
    lines = ["# Directorio ayuntamientos igualdad social", ""]
    coverage = Counter(row["Isla"] for row in rows)
    lines.append("## Resumen")
    lines.append("")
    lines.append(f"- Municipios cubiertos: {len({(row['Isla'], row['Municipio']) for row in rows})}/88")
    lines.append(f"- Filas totales: {len(rows)}")
    lines.append(f"- Contactos recomendados: {sum(1 for row in rows if row['Contacto recomendado'] == 'Si')}")
    confidence = Counter(row["Nivel de confianza"] for row in rows)
    for level in ["Alto", "Medio", "Bajo"]:
        lines.append(f"- Confianza {level.lower()}: {confidence.get(level, 0)}")
    lines.append("")
    lines.append("## Cobertura por isla")
    lines.append("")
    for island in ISLAND_ORDER:
        if island in coverage:
            lines.append(f"- {island}: {coverage[island]} filas")
    lines.append("")
    lines.append("## Tabla")
    lines.append("")
    preview_headers = [
        "Isla",
        "Municipio",
        "Area / Concejalia / Servicio",
        "Cargo exacto",
        "Nombre y apellidos",
        "Email del area / concejalia / servicio",
        "Nivel de confianza",
    ]
    lines.append("| " + " | ".join(preview_headers) + " |")
    lines.append("| " + " | ".join(["---"] * len(preview_headers)) + " |")
    for row in rows:
        values = [row.get(header, "Unknown").replace("\n", " ") for header in preview_headers]
        lines.append("| " + " | ".join(values) + " |")
    DIRECTORY_MD.write_text("\n".join(lines), encoding="utf-8")


def write_coverage_report(rows: list[dict[str, str]]) -> None:
    covered = defaultdict(set)
    recommended = defaultdict(int)
    confidence = defaultdict(Counter)
    for row in rows:
        island = row["Isla"]
        covered[island].add(row["Municipio"])
        if row["Contacto recomendado"] == "Si":
            recommended[island] += 1
        confidence[island][row["Nivel de confianza"]] += 1
    expected = Counter(island for island, *_ in SEED_ROWS)
    lines = ["# Coverage report", "", "| Isla | Municipios esperados | Municipios cubiertos | Filas | Recomendados | Alto | Medio | Bajo |", "| --- | --- | --- | --- | --- | --- | --- | --- |"]
    for island in ISLAND_ORDER:
        island_rows = [row for row in rows if row["Isla"] == island]
        lines.append(
            f"| {island} | {expected[island]} | {len(covered[island])} | {len(island_rows)} | {recommended[island]} | {confidence[island]['Alto']} | {confidence[island]['Medio']} | {confidence[island]['Bajo']} |"
        )
    missing = []
    all_seed = {(row["Isla"], row["Municipio"]) for row in build_seed_rows()}
    all_directory = {(row["Isla"], row["Municipio"]) for row in rows}
    for island, municipality in sorted(all_seed - all_directory):
        missing.append(f"- {island} / {municipality}")
    lines.append("")
    lines.append("## Municipios sin fila en directorio")
    lines.append("")
    lines.extend(missing or ["- Ninguno"])
    COVERAGE_MD.write_text("\n".join(lines), encoding="utf-8")


def write_notes(rows: list[dict[str, str]]) -> None:
    lines = [
        "# Scraping notes ayuntamientos",
        "",
        "Fecha de investigacion: 2026-05-31.",
        "",
        "## Metodo",
        "",
        "- Seed oficial de webs municipales a partir del Gobierno de Canarias.",
        "- Priorizacion de fuentes oficiales municipales: web del ayuntamiento, transparencia, sede, corporacion municipal, concejalias, directorios y noticias institucionales.",
        "- No se inventaron emails ni se dedujeron patrones.",
        "- Cuando no se localizo email personal o de area, se marco `Unknown` y se rebajo la confianza.",
        "",
        "## Distribucion de confianza",
        "",
    ]
    confidence = Counter(row["Nivel de confianza"] for row in rows)
    for level in ["Alto", "Medio", "Bajo"]:
        lines.append(f"- {level}: {confidence.get(level, 0)}")
    lines.append("")
    lines.append("## Gaps tipicos")
    lines.append("")
    lines.append("- Municipios sin email politico personal publicado.")
    lines.append("- Municipios donde el area social o de igualdad aparece solo con telefono o formulario.")
    lines.append("- Webs con delegaciones en noticias o actas, pero sin pagina estructurada de concejalias.")
    NOTES_MD.write_text("\n".join(lines), encoding="utf-8")


def write_checkpoints(rows: list[dict[str, str]]) -> None:
    seed_map = defaultdict(set)
    for row in build_seed_rows():
        seed_map[row["Isla"]].add(row["Municipio"])
    for island in ISLAND_ORDER:
        island_slug = slug(island)
        island_rows = [row for row in rows if row["Isla"] == island]
        lines = [f"# Checkpoint {island}", ""]
        lines.append("## Municipios completados")
        lines.append("")
        covered = sorted({row["Municipio"] for row in island_rows})
        for municipality in covered:
            lines.append(f"- {municipality}")
        lines.append("")
        lines.append("## Fuentes encontradas")
        lines.append("")
        type_counter = Counter(row["Tipo de fuente principal"] for row in island_rows)
        for source_type, count in type_counter.most_common():
            lines.append(f"- {source_type}: {count}")
        lines.append("")
        lines.append("## Problemas tecnicos")
        lines.append("")
        lines.append("- Heterogeneidad de CMS y portales municipales.")
        lines.append("- En varios municipios el email visible pertenece al servicio y no al cargo politico.")
        lines.append("")
        lines.append("## Campos faltantes")
        lines.append("")
        missing_direct_email = sum(1 for row in island_rows if row["Email institucional personal"] == "Unknown")
        missing_area_email = sum(1 for row in island_rows if row["Email del area / concejalia / servicio"] == "Unknown")
        lines.append(f"- Filas sin email personal institucional: {missing_direct_email}")
        lines.append(f"- Filas sin email de area: {missing_area_email}")
        lines.append("")
        lines.append("## Proximos pasos")
        lines.append("")
        if seed_map[island] - set(covered):
            for municipality in sorted(seed_map[island] - set(covered)):
                lines.append(f"- Completar o revisar: {municipality}")
        else:
            lines.append("- Revisar manualmente los casos con confianza baja antes de outreach.")
        (CHECKPOINTS_DIR / f"checkpoint_{island_slug}.md").write_text("\n".join(lines), encoding="utf-8")


def run_seed() -> None:
    ensure_dirs()
    write_csv(SEED_CSV, SEED_HEADERS, build_seed_rows())
    if not RAW_CONTACTS_CSV.exists():
        write_csv(RAW_CONTACTS_CSV, DIRECTORY_HEADERS, [])
    print(f"Seed escrito en {SEED_CSV}")


def run_crawl(args: list[str]) -> None:
    ensure_dirs()
    limit_islands = {value.strip() for value in args[0].split(",")} if args else None
    limit_municipalities = {value.strip() for value in args[1].split(",")} if len(args) > 1 else None
    candidates = crawl_candidates(limit_islands=limit_islands, limit_municipalities=limit_municipalities)
    write_crawl_reports(candidates)
    print(f"Candidatos escritos en {CRAWL_JSON} y {CRAWL_MD}")


def run_build() -> None:
    ensure_dirs()
    rows = read_csv(RAW_CONTACTS_CSV)
    write_csv(DIRECTORY_CSV, DIRECTORY_HEADERS, rows)
    write_xlsx(rows)
    write_markdown(rows)
    write_notes(rows)
    write_coverage_report(rows)
    write_checkpoints(rows)
    with SOURCES_JSON.open("w", encoding="utf-8") as handle:
        json.dump(build_sources_index(rows), handle, ensure_ascii=False, indent=2)
    print(f"Directorio generado en {OUT_DIR}")


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("Uso: python3 create_ayuntamientos_igualdad_social_directory.py [seed|crawl|build] [args...]")
        return 1
    command = argv[1]
    if command == "seed":
        run_seed()
        return 0
    if command == "crawl":
        run_crawl(argv[2:])
        return 0
    if command == "build":
        run_build()
        return 0
    print(f"Comando desconocido: {command}")
    return 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
